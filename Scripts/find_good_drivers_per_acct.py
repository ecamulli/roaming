import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
account_history_dir = "Output/bad_drivers_per_acct"
merged_file_path = "Output/merged_roaming_analysis_with_vintage.csv"
output_dir = "Output/roaming_impact_reports_per_acct"

# Ensure output directory exists
os.makedirs(output_dir, exist_ok=True)

# Load merged analysis
merged_df = pd.read_csv(merged_file_path)
merged_df.columns = [col.strip().lower() for col in merged_df.columns]
merged_df['adapter'] = merged_df['adapter'].str.strip().str.lower()

# Ensure good roaming is numeric
roam_col = 'good roaming calculation (%)'
merged_df[roam_col] = pd.to_numeric(merged_df[roam_col], errors='coerce')

# Loop through account bad driver files
for filename in os.listdir(account_history_dir):
    if filename.startswith("bad_drivers") and filename.endswith(".csv"):
        account_name = filename.replace("bad_drivers_for_", "").replace(".csv", "")
        file_path = os.path.join(account_history_dir, filename)

        bad_df = pd.read_csv(file_path)
        bad_df.columns = [col.strip().lower() for col in bad_df.columns]

        if 'adapter' not in bad_df.columns:
            print(f"[{account_name}] No 'adapter' column. Skipping.")
            continue

        bad_df['adapter'] = bad_df['adapter'].str.strip().str.lower()
        bad_adapters = bad_df['adapter'].unique()

        matched_df = merged_df[merged_df['adapter'].isin(bad_adapters)]
        good_df = matched_df[matched_df[roam_col] > 99.5]

        output_file = os.path.join(output_dir, f"{account_name}_driver_summary.xlsx")

        # Write initial content to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            workbook = writer.book
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            ws = workbook.create_sheet(title='Driver Summary')

            # === WRITE RAW BAD DRIVERS SECTION ===
            ws.append(["Bad Drivers"])
            ws["A" + str(ws.max_row)].font = Font(bold=True)

            bad_driver_start = ws.max_row + 1

            if not bad_df.empty:
                for row in dataframe_to_rows(bad_df, index=False, header=True):
                    ws.append(row)
            else:
                ws.append(["No bad drivers found."])

        # Reopen to clean and format
        wb = load_workbook(output_file)
        ws = wb.active

        # === CLEAN & REWRITE BAD DRIVERS SECTION ===
        start_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_col=10), start=1):
            if row[0].value == "Bad Drivers":
                start_row = i + 1
                break

        if start_row is not None:
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=start_row, max_row=start_row))]
            rows_to_keep = []
            for r in ws.iter_rows(min_row=start_row + 1, max_col=10, values_only=True):
                if all(v is None for v in r):
                    break
                data = dict(zip(header_row, r))
                rows_to_keep.append([
                    data.get('adapter-driver'),
                    data.get('client count'),
                    data.get('critical minutes'),
                    data.get('good roaming calculation (%)')
                ])

            rows_to_clear = len(rows_to_keep) + 5
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + rows_to_clear):
                for cell in row:
                    cell.value = None

            cleaned_headers = ['Adapter-Driver', 'Client Count', 'Critical Minutes', 'Good Roaming Calculation (%)']
            for col, header in enumerate(cleaned_headers, start=1):
                ws.cell(row=start_row, column=col, value=header)

            # Underline the Bad Drivers header row
            thin_border = Border(bottom=Side(style='thin'))
            for col in range(1, len(cleaned_headers) + 1):
                ws.cell(row=start_row, column=col).border = thin_border

            for i, row_data in enumerate(rows_to_keep):
                for j, value in enumerate(row_data, start=1):
                    ws.cell(row=start_row + 1 + i, column=j, value=value)

            total_row_index = start_row + 1 + len(rows_to_keep)
            ws.cell(row=total_row_index, column=1, value="Totals:")
            ws.cell(row=total_row_index, column=2, value=sum(val[1] for val in rows_to_keep if isinstance(val[1], (int, float))))
            ws.cell(row=total_row_index, column=3, value=sum(val[2] for val in rows_to_keep if isinstance(val[2], (int, float))))
            ws.cell(row=total_row_index, column=1).font = Font(bold=True)
            ws.cell(row=total_row_index, column=2).font = Font(bold=True)
            ws.cell(row=total_row_index, column=3).font = Font(bold=True)

            # Left-align all cells in column B
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep), min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')

            # Left-align all cells in column C
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep), min_col=3, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')


            # Left-align all cells in column D
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep), min_col=4, max_col=4):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')

            for col in [2, 3]:
                cell = ws.cell(row=total_row_index, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

            ws.append([])
            

        # === WRITE GOOD DRIVERS SECTION ===
        ws.append(["Good Drivers (Roaming > 99.8%)"])
        ws["A" + str(ws.max_row)].font = Font(bold=True)

        if not good_df.empty:
            for row in dataframe_to_rows(good_df, index=False, header=True):
                ws.append(row)
        else:
            ws.append(["No good drivers found."])

        # === FORMAT GOOD DRIVERS SECTION ===
        start_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_col=10), start=1):
            if row[0].value == "Good Drivers (Roaming > 99.5%)":
                start_row = i + 1
                break

        if start_row is not None:
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=start_row, max_row=start_row))]
            rows_to_keep = []
            for r in ws.iter_rows(min_row=start_row + 1, max_col=10, values_only=True):
                if all(v is None for v in r):
                    break
                data = dict(zip(header_row, r))
                rows_to_keep.append([
                    data.get('adapter-driver'),
                    data.get('total sum'),
                    '',
                    data.get('good roaming calculation (%)'),
                    data.get('driver vintage')
                ])

            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep) + 5):
                for cell in row:
                    cell.value = None

            new_headers = ['Adapter-Driver', 'Total Samples', '', 'Good Roaming Calculation (%)', 'Driver Vintage']
            for col, header in enumerate(new_headers, start=1):
                ws.cell(row=start_row, column=col, value=header)

            
            driver_vintage_column = new_headers.index("Driver Vintage") + 1
            rows_to_keep.sort(
                key=lambda row: str(row[driver_vintage_column - 1]) if row[driver_vintage_column - 1] is not None else '',
                reverse=True
            )
            for i, row_data in enumerate(rows_to_keep, start=start_row + 1):
                for j, value in enumerate(row_data, start=1):
                    ws.cell(row=i, column=j, value=value)

            # Underline the Good Drivers header row
            thin_border = Border(bottom=Side(style='thin'))
            for col in range(1, len(new_headers) + 1):
                ws.cell(row=start_row, column=col).border = thin_border

            # Left-align all cells in column B
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep), min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')


            # Left-align all cells in column D
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep), min_col=4, max_col=4):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')


            # Left-align all cells in column E
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows_to_keep), min_col=5, max_col=5):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')

            for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + len(rows_to_keep), min_col=2, max_col=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

        # === AUTO-SIZE COLUMNS ===
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            column_letter = get_column_letter(column)
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(output_file)
        print(f"✅ {account_name} report generated and formatted.")
